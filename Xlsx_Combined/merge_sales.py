'''
매출장 파일에서 '세금계산서'라는 단어가 들어간 시트를 검색
검색 된 시트를 '파일명.csv'라는 파일로 저장하는 프로그램
'''

import openpyxl
import csv
import time

def process_excel_files():
    total_files = 12
    start_time = time.time()
    last_update_time = start_time
    
    for month in range(1, 13):
        input_filename = f'2024{month:02d}.xlsx'
        output_filename = f'2024{month:02d}.csv'
        try:
            wb = openpyxl.load_workbook(input_filename, data_only=True)
            
            tax_invoice_sheet = None
            for sheet_name in wb.sheetnames:
                if '세금계산서' in sheet_name:
                    tax_invoice_sheet = wb[sheet_name]
                    break
            
            if tax_invoice_sheet:
                with open(output_filename, 'w', newline='', encoding='ansi') as csvfile:
                    csv_writer = csv.writer(csvfile)
                    for row in tax_invoice_sheet.iter_rows(values_only=True):
                        csv_writer.writerow(row)
            
            progress = (month / total_files) * 100
            current_time = time.time()
            if current_time - last_update_time >= 5:
                elapsed_time = current_time - start_time
                estimated_total_time = (elapsed_time / month) * total_files
                remaining_time = estimated_total_time - elapsed_time
                print(f"\r진행률: {progress:.2f}% | 예상 남은 시간: {remaining_time:.2f}초", end='')
                last_update_time = current_time
        
        except FileNotFoundError:
            print(f"\n{input_filename} 파일을 찾을 수 없습니다.")
        except Exception as e:
            print(f"\n{input_filename} 처리 중 오류 발생: {str(e)}")
    
    print("\n모든 작업이 완료되었습니다. 각 파일이 개별 CSV 파일로 저장되었습니다.")

process_excel_files()
