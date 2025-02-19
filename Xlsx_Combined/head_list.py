'''
매출장 파일에서 '세금계산서'라는 단어가 들어간 시트를 검색
검색 된 시트의 헤더(1행) 값을 출력해주는 프로그램
'''

import openpyxl
import time

def process_excel_files():
    total_files = 12
    start_time = time.time()
    
    with open('head_list.txt', 'w') as output_file:
        for month in range(1, 13):
            filename = f'2024{month:02d}.xlsx'
            try:
                workbook = openpyxl.load_workbook(filename)
                
                tax_invoice_sheet = None
                for sheet_name in workbook.sheetnames:
                    if '세금계산서' in sheet_name:
                        tax_invoice_sheet = workbook[sheet_name]
                        break
                
                if tax_invoice_sheet:
                    header = [tax_invoice_sheet.cell(row=1, column=col).value for col in range(1, tax_invoice_sheet.max_column + 1)]
                    header_str = ' | '.join(str(cell) if cell is not None else '' for cell in header)
                    output_file.write(f'{filename[:6]} - {header_str}\n')
                else:
                    output_file.write(f'{filename[:6]} - 세금계산서 시트를 찾을 수 없습니다.\n')
            
            except FileNotFoundError:
                output_file.write(f'{filename[:6]} - 파일을 찾을 수 없습니다.\n')
            except Exception as e:
                output_file.write(f'{filename[:6]} - 오류 발생: {str(e)}\n')
            
            progress = (month / total_files) * 100
            elapsed_time = time.time() - start_time
            estimated_total_time = (elapsed_time / month) * total_files
            remaining_time = estimated_total_time - elapsed_time
            
            print(f"\r작업 진행률: {progress:.2f}% | 예상 남은 시간: {remaining_time:.2f}초", end='')
    
    print("\n모든 작업이 완료되었습니다. 결과가 head_list.txt 파일에 저장되었습니다.")

process_excel_files()
